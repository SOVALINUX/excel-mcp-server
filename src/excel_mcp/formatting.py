import logging
import re
from datetime import datetime, date
from typing import Any, Dict, Optional

from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Protection, Font,
    Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    FormulaRule, CellIsRule
)
from openpyxl.utils import get_column_letter

from .workbook import get_or_create_workbook
from .cell_utils import parse_cell_range, validate_cell_reference, get_actual_data_range
from .exceptions import ValidationError, FormattingError

logger = logging.getLogger(__name__)


def _parse_date_string(value: str, known_format: str | None = None) -> tuple[datetime | None, str | None]:
    """Try to parse a string as a date in common formats.
    
    Supports:
    - ISO format: 2024-01-15, 2024-01-15T10:30:00
    - US format: 01/15/2024, 1/15/2024
    - EU format: 15/01/2024, 15.01.2024
    - Text format: 15-Jan-2024, Jan 15 2024
    
    Args:
        value: String value to parse
        known_format: Optional format string from previous successful parse in same column.
                     If provided, this format is tried first for performance.
        
    Returns:
        Tuple of (datetime object if parsing succeeds, format string used)
        Returns (None, None) if parsing fails
    """
    if not isinstance(value, str):
        return None, None
    
    # Clean the string
    value = value.strip()
    
    # If we know the format from previous cell, try it first
    if known_format:
        try:
            return datetime.strptime(value, known_format), known_format
        except ValueError:
            # Format changed or invalid, fall through to try all formats
            pass
    
    # Common date formats to try
    date_formats = [
        # ISO formats with milliseconds
        '%Y-%m-%d %H:%M:%S.%f',
        '%Y-%m-%dT%H:%M:%S.%f',
        # ISO formats
        '%Y-%m-%d',
        '%Y-%m-%dT%H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%Y/%m/%d',
        '%Y/%m/%d %H:%M:%S',
        # US formats
        '%m/%d/%Y',
        '%m-%d-%Y',
        '%m.%d.%Y',
        '%m/%d/%Y %H:%M:%S',
        # EU formats
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%d.%m.%Y',
        '%d/%m/%Y %H:%M:%S',
        # Text formats
        '%d-%b-%Y',
        '%b %d, %Y',
        '%B %d, %Y',
        '%d %b %Y',
        '%d %B %Y',
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(value, fmt), fmt
        except ValueError:
            continue
    
    return None, None


def _is_date_like(value: Any, known_format: str | None = None) -> tuple[bool, str | None]:
    """Check if a value is a date or looks like a date string.
    
    Args:
        value: Value to check
        known_format: Optional format string from previous successful parse in same column
        
    Returns:
        Tuple of (True if value is a date/datetime/parseable date string, format string if parsed)
    """
    if isinstance(value, (datetime, date)):
        return True, None
    
    if isinstance(value, str):
        parsed_date, fmt = _parse_date_string(value, known_format)
        return parsed_date is not None, fmt
    
    return False, None


def _count_significant_digits(value: str) -> int:
    """Count significant digits in a numeric string.
    
    Excel stores numbers as IEEE-754 double-precision floats with 15 significant digits.
    Numbers with more than 15 digits will lose precision.
    
    Args:
        value: String representation of a number
        
    Returns:
        Number of significant digits (excluding leading zeros, decimal point, signs)
        
    Examples:
        '123' -> 3
        '0012.3400' -> 5 (leading zeros don't count, trailing zeros after decimal do)
        '1234567890123456' -> 16
        '-123.456' -> 6
    """
    # Remove whitespace and sign
    cleaned = value.strip().lstrip('+-')
    
    # Split by decimal point
    if '.' in cleaned:
        integer_part, decimal_part = cleaned.split('.', 1)
        # Remove leading zeros from integer part
        integer_part = integer_part.lstrip('0')
        # Keep trailing zeros in decimal part as they are significant
        # But remove trailing zeros for counting
        decimal_part = decimal_part.rstrip('0')
        # Combine and count
        significant = integer_part + decimal_part
        if not significant:
            return 1
        return len(significant)
    else:
        # No decimal point - remove commas and leading zeros
        cleaned = cleaned.replace(',', '').lstrip('0')
        if not cleaned:
            return 1
        return len(cleaned)

def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
    auto_column_width: bool = False,
    column_width: Optional[float] = None,
    auto_detect_numeric_columns: bool = False,
    date_format: Optional[str] = None,
    auto_detect_date_columns: bool = False
) -> Dict[str, Any]:
    """Apply formatting to a range of cells.
    
    This function handles all Excel formatting operations including:
    - Font properties (bold, italic, size, color, etc.)
    - Cell fill/background color
    - Borders (style and color)
    - Number formatting
    - Alignment and text wrapping
    - Cell merging
    - Protection
    - Conditional formatting
    - Auto column width detection
    - Auto-detection of numeric and date columns
    
    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell reference
        end_cell: Optional ending cell reference
        bold: Whether to make text bold
        italic: Whether to make text italic
        underline: Whether to underline text
        font_size: Font size in points
        font_color: Font color (hex code)
        bg_color: Background color (hex code)
        border_style: Border style (thin, medium, thick, double)
        border_color: Border color (hex code)
        number_format: Excel number format string
        alignment: Text alignment (left, center, right, justify)
        wrap_text: Whether to wrap text
        merge_cells: Whether to merge the range
        protection: Cell protection settings
        conditional_format: Conditional formatting rules
        auto_column_width: Auto-adjust column width based on content (approximate)
        column_width: Absolute column width (applied to all columns in range)
        auto_detect_numeric_columns: Auto-detect and apply number format to numeric columns
        date_format: Date format string (e.g., 'yyyy-mm-dd')
        auto_detect_date_columns: Auto-detect and apply date format to date columns
        
    Returns:
        Dictionary with operation status
    """
    try:
        # Validate cell references
        if not validate_cell_reference(start_cell):
            raise ValidationError(f"Invalid start cell reference: {start_cell}")
            
        if end_cell and not validate_cell_reference(end_cell):
            raise ValidationError(f"Invalid end cell reference: {end_cell}")
            
        wb = get_or_create_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"Sheet '{sheet_name}' not found")
            
        sheet = wb[sheet_name]
        
        # Get cell range coordinates
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise ValidationError(f"Invalid cell range: {str(e)}")
        
        # If no end cell specified, use start cell coordinates
        if end_row is None:
            end_row = start_row
        if end_col is None:
            end_col = start_col
            
        # Apply font formatting
        font_args = {
            "bold": bold,
            "italic": italic,
            "underline": 'single' if underline else None,
        }
        if font_size is not None:
            font_args["size"] = font_size
        if font_color is not None:
            try:
                # Ensure color has FF prefix for full opacity
                font_color = font_color if font_color.startswith('FF') else f'FF{font_color}'
                font_args["color"] = Color(rgb=font_color)
            except ValueError as e:
                raise FormattingError(f"Invalid font color: {str(e)}")
        font = Font(**font_args)
        
        # Apply fill
        fill = None
        if bg_color is not None:
            try:
                # Ensure color has FF prefix for full opacity
                bg_color = bg_color if bg_color.startswith('FF') else f'FF{bg_color}'
                fill = PatternFill(
                    start_color=Color(rgb=bg_color),
                    end_color=Color(rgb=bg_color),
                    fill_type='solid'
                )
            except ValueError as e:
                raise FormattingError(f"Invalid background color: {str(e)}")
        
        # Apply borders
        border = None
        if border_style is not None:
            try:
                border_color = border_color if border_color else "000000"
                border_color = border_color if border_color.startswith('FF') else f'FF{border_color}'
                side = Side(
                    style=border_style,
                    color=Color(rgb=border_color)
                )
                border = Border(
                    left=side,
                    right=side,
                    top=side,
                    bottom=side
                )
            except ValueError as e:
                raise FormattingError(f"Invalid border settings: {str(e)}")
            
        # Apply alignment
        align = None
        if alignment is not None or wrap_text:
            try:
                align = Alignment(
                    horizontal=alignment,
                    vertical='center',
                    wrap_text=wrap_text
                )
            except ValueError as e:
                raise FormattingError(f"Invalid alignment settings: {str(e)}")
            
        # Apply protection
        protect = None
        if protection is not None:
            try:
                protect = Protection(**protection)
            except ValueError as e:
                raise FormattingError(f"Invalid protection settings: {str(e)}")
            
        # Apply formatting to range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = font
                if fill is not None:
                    cell.fill = fill
                if border is not None:
                    cell.border = border
                if align is not None:
                    cell.alignment = align
                if protect is not None:
                    cell.protection = protect
                if number_format is not None:
                    cell.number_format = number_format
                
                # Apply date format if specified (convert string dates to datetime)
                if date_format is not None and not auto_detect_date_columns:
                    if isinstance(cell.value, str):
                        parsed_date, _ = _parse_date_string(cell.value)
                        if parsed_date:
                            cell.value = parsed_date
                    cell.number_format = date_format
                    
        # Merge cells if requested
        if merge_cells and end_cell:
            try:
                range_str = f"{start_cell}:{end_cell}"
                sheet.merge_cells(range_str)
            except ValueError as e:
                raise FormattingError(f"Failed to merge cells: {str(e)}")
            
        # Apply conditional formatting
        if conditional_format is not None:
            range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
            rule_type = conditional_format.get('type')
            if not rule_type:
                raise FormattingError("Conditional format type not specified")
                
            params = conditional_format.get('params', {})
            
            # Handle fill parameter for cell_is rule
            if rule_type == 'cell_is' and 'fill' in params:
                fill_params = params['fill']
                if isinstance(fill_params, dict):
                    try:
                        fill_color = fill_params.get('fgColor', 'FFC7CE')  # Default to light red
                        fill_color = fill_color if fill_color.startswith('FF') else f'FF{fill_color}'
                        params['fill'] = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type='solid'
                        )
                    except ValueError as e:
                        raise FormattingError(f"Invalid conditional format fill color: {str(e)}")
            
            try:
                if rule_type == 'color_scale':
                    rule = ColorScaleRule(**params)
                elif rule_type == 'data_bar':
                    rule = DataBarRule(**params)
                elif rule_type == 'icon_set':
                    rule = IconSetRule(**params)
                elif rule_type == 'formula':
                    rule = FormulaRule(**params)
                elif rule_type == 'cell_is':
                    rule = CellIsRule(**params)
                else:
                    raise FormattingError(f"Invalid conditional format type: {rule_type}")
                    
                sheet.conditional_formatting.add(range_str, rule)
            except Exception as e:
                raise FormattingError(f"Failed to apply conditional formatting: {str(e)}")
        
        # Track auto-detection results
        auto_detection_results = {
            "numeric_columns": [],
            "date_columns": [],
            "datetime_columns": [],
            "text_columns": [],
            "long_number_columns": []
        }
        
        # Auto-detect and apply formats to columns
        # OPTIMIZATION: Only scan rows that actually have data
        if auto_detect_numeric_columns or auto_detect_date_columns:
            # Find the actual data range (stops early at empty rows)
            max_data_row, max_data_col = get_actual_data_range(
                sheet, start_row, start_col, end_row, end_col, max_empty_rows=10
            )
            
            # Analyze column data types (only scan rows with data)
            for col in range(start_col, end_col + 1):
                is_numeric = True
                is_date = True
                is_datetime = False  # Track if we have datetime vs pure date
                has_data = False
                detected_date_format = None  # Cache the date format for this column
                has_long_number = False  # Track if column has numbers > 15 digits
                
                # Excel's precision limit: 15 significant digits (IEEE-754 double precision)
                EXCEL_MAX_PRECISION_DIGITS = 15
                
                # Only scan up to max_data_row instead of entire range
                for row in range(start_row, min(max_data_row + 1, end_row + 1)):
                    cell = sheet.cell(row=row, column=col)
                    value = cell.value
                    
                    if value is not None and value != '':
                        has_data = True
                        
                        # Check if numeric (including string representations)
                        if isinstance(value, (int, float)):
                            # Already numeric, not a date
                            is_date = False
                            # Check if it's too long (would lose precision)
                            # Convert to string to count digits
                            str_value = str(value)
                            if _count_significant_digits(str_value) > EXCEL_MAX_PRECISION_DIGITS:
                                has_long_number = True
                                is_numeric = False
                        elif isinstance(value, str):
                            # Try to parse as number
                            try:
                                float(value)
                                # It's a numeric string, not a date
                                is_date = False
                                # Check if it's too long for Excel's precision
                                if _count_significant_digits(value) > EXCEL_MAX_PRECISION_DIGITS:
                                    has_long_number = True
                                    is_numeric = False
                            except ValueError:
                                # Not numeric, check if it's a date
                                is_numeric = False
                                is_date_value, date_fmt = _is_date_like(value, detected_date_format)
                                if not is_date_value:
                                    is_date = False
                                else:
                                    # Cache the format for subsequent rows
                                    if date_fmt:
                                        detected_date_format = date_fmt
                                    # Check if it has time component (datetime vs date)
                                    if ' ' in value or 'T' in value:
                                        is_datetime = True
                        else:
                            # Check if date object
                            if isinstance(value, datetime):
                                is_numeric = False
                                is_datetime = True
                            elif isinstance(value, date):
                                is_numeric = False
                            else:
                                is_numeric = False
                                is_date = False
                        
                        # Early exit if both checks failed
                        if not is_numeric and not is_date:
                            break
                
                # Apply formats based on detection and track results
                col_letter = get_column_letter(col)
                if has_data:
                    # Log if column has long numbers that are kept as text
                    if has_long_number:
                        logger.info(
                            f"Column {col_letter} contains numbers with >15 significant digits. "
                            f"Keeping as text to prevent data loss (Excel's precision limit is 15 digits)."
                        )
                        auto_detection_results["long_number_columns"].append(col_letter)
                    
                    if auto_detect_numeric_columns and is_numeric and number_format:
                        # Convert string numbers to actual numbers and apply format
                        for row in range(start_row, end_row + 1):
                            cell = sheet.cell(row=row, column=col)
                            if isinstance(cell.value, str) and cell.value.strip():
                                try:
                                    # Try integer first, then float
                                    if '.' not in cell.value:
                                        cell.value = int(cell.value)
                                    else:
                                        cell.value = float(cell.value)
                                except ValueError:
                                    pass  # Keep as string if conversion fails
                            cell.number_format = number_format
                        auto_detection_results["numeric_columns"].append(col_letter)
                    
                    if auto_detect_date_columns and is_date:
                        # Determine appropriate date format
                        if date_format:
                            actual_date_format = date_format
                        elif is_datetime:
                            # Default datetime format
                            actual_date_format = 'yyyy-mm-dd hh:mm:ss'
                        else:
                            # Default date-only format
                            actual_date_format = 'yyyy-mm-dd'
                        
                        # Convert date strings to Excel date objects and apply format
                        # Use cached format for performance
                        column_date_format = detected_date_format
                        for row in range(start_row, end_row + 1):
                            cell = sheet.cell(row=row, column=col)
                            
                            # Convert string dates to datetime objects
                            if isinstance(cell.value, str) and cell.value.strip():
                                parsed_date, fmt = _parse_date_string(cell.value, column_date_format)
                                if parsed_date:
                                    cell.value = parsed_date
                                    # Cache format for next iteration
                                    if fmt and not column_date_format:
                                        column_date_format = fmt
                            
                            # Apply date format
                            cell.number_format = actual_date_format
                        
                        # Track whether it's datetime or date
                        if is_datetime:
                            auto_detection_results["datetime_columns"].append(col_letter)
                        else:
                            auto_detection_results["date_columns"].append(col_letter)
                    
                    # Track text columns (columns that weren't detected as numeric or date)
                    if not is_numeric and not is_date and not has_long_number:
                        auto_detection_results["text_columns"].append(col_letter)
        
        # Apply column width settings
        if column_width is not None:
            # Apply absolute width to all columns in range
            for col in range(start_col, end_col + 1):
                col_letter = get_column_letter(col)
                sheet.column_dimensions[col_letter].width = column_width
        
        if auto_column_width:
            # Auto-adjust width based on content
            # OPTIMIZATION: Only scan rows with actual data
            # Reuse data range if already calculated, otherwise calculate it
            if auto_detect_numeric_columns or auto_detect_date_columns:
                # Already calculated above
                scan_end_row = max_data_row
            else:
                # Calculate data range for width calculation only
                scan_end_row, _ = get_actual_data_range(
                    sheet, start_row, start_col, end_row, end_col, max_empty_rows=10
                )
            
            for col in range(start_col, end_col + 1):
                max_length = 0
                col_letter = get_column_letter(col)
                
                # ALWAYS check row 1 (header row) first, even if not in the formatting range
                # This ensures column width accommodates the header
                header_cell = sheet.cell(row=1, column=col)
                if header_cell.value:
                    header_str = str(header_cell.value)
                    header_lines = header_str.split('\n')
                    header_length = max(len(line) for line in header_lines)
                    max_length = header_length
                
                # Then scan the data rows (only up to the last row with data)
                for row in range(start_row, min(scan_end_row + 1, end_row + 1)):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value:
                        # Convert to string and handle multi-line content
                        cell_str = str(cell.value)
                        lines = cell_str.split('\n')
                        # Get the longest line
                        line_length = max(len(line) for line in lines)
                        max_length = max(max_length, line_length)
                
                # Approximate width calculation (character width ~1.2)
                # Add padding for better appearance
                if max_length > 0:
                    adjusted_width = (max_length * 1.2) + 2
                    # Cap maximum width to 100 for readability
                    adjusted_width = min(adjusted_width, 100)
                    sheet.column_dimensions[col_letter].width = adjusted_width
            
        wb.save(filepath)
        
        range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
        result = {
            "message": f"Applied formatting to range {range_str}",
            "range": range_str
        }
        
        # Add auto-detection results if any detection was performed
        if auto_detect_numeric_columns or auto_detect_date_columns:
            result["auto_detection"] = {
                "numeric_columns": auto_detection_results["numeric_columns"],
                "date_columns": auto_detection_results["date_columns"],
                "datetime_columns": auto_detection_results["datetime_columns"],
                "text_columns": auto_detection_results["text_columns"],
                "long_number_columns": auto_detection_results["long_number_columns"]
            }
            
            # Add summary message with column lists
            summary_parts = []
            if auto_detection_results["numeric_columns"]:
                cols = ", ".join(auto_detection_results['numeric_columns'])
                summary_parts.append(f"numeric: {cols}")
            if auto_detection_results["date_columns"]:
                cols = ", ".join(auto_detection_results['date_columns'])
                summary_parts.append(f"date: {cols}")
            if auto_detection_results["datetime_columns"]:
                cols = ", ".join(auto_detection_results['datetime_columns'])
                summary_parts.append(f"datetime: {cols}")
            if auto_detection_results["long_number_columns"]:
                cols = ", ".join(auto_detection_results['long_number_columns'])
                summary_parts.append(f"long number (kept as text): {cols}")
            if auto_detection_results["text_columns"]:
                cols = ", ".join(auto_detection_results['text_columns'])
                summary_parts.append(f"text: {cols}")
            
            if summary_parts:
                result["auto_detection"]["summary"] = f"Detected columns - {'; '.join(summary_parts)}"
        
        return result
        
    except (ValidationError, FormattingError) as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to apply formatting: {e}")
        raise FormattingError(str(e))
