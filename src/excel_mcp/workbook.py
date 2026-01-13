import base64
import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .exceptions import WorkbookError

logger = logging.getLogger(__name__)

def create_workbook(filepath: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name"""
    try:
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
            "workbook": wb
        }
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        raise WorkbookError(f"Failed to create workbook: {e!s}")

def get_or_create_workbook(filepath: str) -> Workbook:
    """Get existing workbook or create new one if it doesn't exist"""
    try:
        return load_workbook(filepath)
    except FileNotFoundError:
        return create_workbook(filepath)["workbook"]

def create_sheet(filepath: str, sheet_name: str) -> dict:
    """Create a new worksheet in the workbook if it doesn't exist."""
    try:
        wb = load_workbook(filepath)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            raise WorkbookError(f"Sheet {sheet_name} already exists")

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(filepath)
        wb.close()
        return {"message": f"Sheet {sheet_name} created successfully"}
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to create sheet: {e}")
        raise WorkbookError(str(e))

def get_workbook_info(filepath: str, include_ranges: bool = False) -> dict[str, Any]:
    """Get metadata about workbook including sheets, ranges, etc."""
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
            
        wb = load_workbook(filepath, read_only=False)
        
        info = {
            "filename": path.name,
            "sheets": wb.sheetnames,
            "size": path.stat().st_size,
            "modified": path.stat().st_mtime
        }
        
        if include_ranges:
            # Add used ranges for each sheet
            ranges = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.max_row > 0 and ws.max_column > 0:
                    ranges[sheet_name] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            info["used_ranges"] = ranges
            
        wb.close()
        return info
        
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Failed to get workbook info: {e}")
        raise WorkbookError(str(e))

def read_excel_binary(filepath: str) -> str:
    """Read an Excel file and return its binary content as a base64-encoded string.
    
    This function is useful when you need to:
    - Transfer Excel files through text-based protocols
    - Upload Excel files to cloud storage (S3, Azure Blob, Google Cloud Storage, etc.)
    - Send Excel files through APIs that accept base64-encoded data
    - Embed Excel files in JSON or other text-based formats
    - Store Excel files in databases as text fields
    
    The returned base64 string can be decoded back to binary format and written
    to a file or uploaded to cloud storage services.
    
    Args:
        filepath: Path to the Excel file to read (supports .xlsx, .xlsm, .xlsb, .xls formats)
    
    Returns:
        Base64-encoded string of the Excel file binary content
    
    Raises:
        WorkbookError: If file not found, permission denied, or other read errors
    
    Example:
        base64_content = read_excel_binary(filepath='reports/sales_report.xlsx')
        
        # The base64 content can now be:
        # - Sent to an API endpoint
        # - Uploaded to AWS S3, Azure Blob Storage, etc.
        # - Transferred over a network
        # - Stored in a database
        # - Embedded in a JSON payload
    
    Notes:
        - The file must exist and be readable
        - Works with .xlsx, .xlsm, .xlsb (Excel 2007+) and .xls (Excel 97-2003) formats
        - The base64 string size will be approximately 33% larger than the original file
        - For very large files (>50MB), consider streaming or chunking approaches
        - Use get_workbook_metadata() to retrieve file metadata (size, sheets, etc.)
        - To decode in Python: base64.b64decode(content)
        - To decode in Node.js: Buffer.from(content, 'base64')
        - To decode in browser JavaScript: atob(content)
    """
    try:
        path = Path(filepath)
        if not path.exists():
            raise WorkbookError(f"File not found: {filepath}")
        
        # Get file size for logging
        file_size = path.stat().st_size
        
        # Read the file as binary
        with open(filepath, 'rb') as f:
            binary_content = f.read()
        
        # Encode to base64
        base64_content = base64.b64encode(binary_content).decode('utf-8')
        
        logger.info(f"Successfully read Excel file as binary: {filepath} ({file_size} bytes, base64 size: {len(base64_content)} chars)")
        
        return base64_content
        
    except PermissionError:
        error_msg = f"Permission denied reading file: {filepath}"
        logger.error(error_msg)
        raise WorkbookError(error_msg)
    except WorkbookError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        error_msg = f"Failed to read Excel file as binary: {str(e)}"
        logger.error(error_msg)
        raise WorkbookError(error_msg)


